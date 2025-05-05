import jinja2
import pdfkit
import openpyxl
from openpyxl import load_workbook
import os
import requests
import win32com.client
import os

nameExcel="CDA_certifies-4-participantWorkshops.xlsx"

def open_excel_file(filepath):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    wb = excel.Workbooks.Open(os.path.abspath(filepath))
    return wb, excel

def obtener_archivos_excel_abiertos():
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        nombres = [wb.Name for wb in excel.Workbooks]
        return nombres
    except Exception as e:
        print("Excel "+nameExcel+" cerrado, el programa puede ejecutarse correctamente")
        return []

# Ejecutar
archivos_abiertos = obtener_archivos_excel_abiertos()

abierto=False
for nombre in archivos_abiertos:
    if nombre==nameExcel:
        abierto=True

# Limpiar consola
os.system('cls' if os.name == 'nt' else 'clear')

Server = 'http://127.0.0.1:5500/'
html_template = 'congratulations.html'

def pdf(ID):
        
            context = {'NameAuthor': Nombres,
                    'Investigation': Investigacion, 'Year': Year, 'InitDate': Fecha_inicio,
                        'FinishDate': Fecha_finalizacion, 'Month':Mes,'ID':ID,'Server':Server,      
            }

            options = {
                'page-width': '215.9mm',   # Ancho de página (carta en pulgadas)
                'page-height': '279.4mm',
            'margin-top': '10mm',   # Ajusta los márgenes superiores según sea necesario
                'margin-right': '5mm',  # Ajusta los márgenes derechos según sea necesario
                'margin-bottom': '0mm',  # Ajusta los márgenes inferiores según sea necesario
                'margin-left': '0mm',   'print-media-type': None ,
                }
            output_text =jinja2.Environment(loader=jinja2.FileSystemLoader(searchpath="./")).get_template(html_template).render(context)
            config = pdfkit.configuration(wkhtmltopdf='wkhtmltopdf/bin/wkhtmltopdf.exe')
            output_pdf = 'Congratulations'+str(ID)+'.pdf'
            pdfkit.from_string(output_text, output_pdf, configuration=config, css='congratulations.css',options=options)

try:
    response = requests.get(Server, timeout=5)  # timeout evita que se quede esperando demasiado
    if response.status_code == 200:
        if(abierto==False):
            #EXCEL
            # Define variable to load the dataframe
            excel_dataframe = openpyxl.load_workbook("CDA_certifies-4-participantWorkshops.xlsx")

            # Define variable to read sheet
            dataframe = excel_dataframe.active
                
            # Cargar el archivo Excel
            wb = load_workbook('CDA_certifies-4-participantWorkshops.xlsx', read_only=True)

            # Obtener una lista de los nombres de las hojas
            sheet_names = wb.sheetnames

            # Iterar sobre cada hoja
            for sheet_name in sheet_names:
                # Obtener la hoja de trabajo por su nombre
                sheet = wb[sheet_name]
                
                # Iterar sobre las filas y columnas de la hoja
                for row in sheet.iter_rows(values_only=True):
                    _row = [row,]
                    for value in row:
                        _row.append(value)
                    
                    if(sheet_name=="Sheet1"):
                        
                        if(row[0]=="General Chair"):
                        
                            Mes=_row[5]
                            Fecha_inicio=_row[6]
                        elif(row[0]=="Jefa del Departamento de Investigación" or row[0]=="Director"):
                        
                            Fecha_finalizacion=_row[6]
                            Year=_row[7]
                                                
                    if sheet_name=="Sheet2":
                        if(row[0]=="ID"):
                            print("Congratulations") 
                        else:
                            ID=_row[1]
                            Investigacion=_row[2]
                            Nombres=_row[3]
                            if(ID!= None):
                                pdf(ID)
                                print("PDF "+ID+" "+Nombres+" ready")
                                
                # Archivos a manejar
            archivo_a_abrir = "C:\\Users\\tona2\\OneDrive\\Desktop\\Certifies_2024\\"+nameExcel
            # Abrir archivo cuando estamos haciendo pruebas
            print("Abriendo archivo "+nameExcel)
            open_excel_file(archivo_a_abrir)
        else:
            print("Excel "+nameExcel+" abierto, debes cerrarlo para poder continuar")
         
    else:
        print(f"El servidor respondió con un código de estado: {response.status_code}")
except requests.exceptions.RequestException as e:
    print("El servidor no está disponible.\nError:", e)