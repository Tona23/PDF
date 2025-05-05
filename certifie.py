import jinja2
import pdfkit
import openpyxl
from openpyxl import load_workbook
import os
import requests
import win32com.client
import os

nameExcel="CDA_certifies-4-participantWorkshops.xlsx"
# Limpiar consola
os.system('cls' if os.name == 'nt' else 'clear')

# Definir la URL de la imagen (puedes modificarla dinámicamente)
html_template = 'certifie.html'
server="http://127.0.0.1:5500/" 

def pdf(nombre):

    options = {
        'page-width': '215.9mm',   # Ancho de página (carta en pulgadas)
        'page-height': '279.4mm',
    'margin-top': '0mm',   # Ajusta los márgenes superiores según sea necesario
        'margin-right': '0mm',  # Ajusta los márgenes derechos según sea necesario
        'margin-bottom': '0mm',  # Ajusta los márgenes inferiores según sea necesario
        'margin-left': '0mm',   'print-media-type': None ,
        }
    
    output_text =jinja2.Environment(loader=jinja2.FileSystemLoader(searchpath="./")).get_template(html_template).render("")
    config = pdfkit.configuration(wkhtmltopdf='wkhtmltopdf/bin/wkhtmltopdf.exe')
    output_pdf = 'Certifie '+str(row[0])+' '+str(row[4])+' '+nombre+'.pdf'
    pdfkit.from_string(output_text, output_pdf, configuration=config, css='certifie.css',options=options)

def open_excel_file(filepath):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    wb = excel.Workbooks.Open(os.path.abspath(filepath))
    return wb, excel

def obtener_archivos_excel_abiertos():
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        nombres = [wb.Name for wb in excel.Workbooks]
        return nombres
    except Exception as e:
        print("Excel "+nameExcel+" cerrado, el programa puede ejecutarse correctamente")
        return []

try:
    response = requests.get(server, timeout=5)  # timeout evita que se quede esperando demasiado
    if response.status_code == 200:
        # Ejecutar
        archivos_abiertos = obtener_archivos_excel_abiertos()

        abierto=False
        for nombre in archivos_abiertos:
            if nombre==nameExcel:
                abierto=True

        if(abierto==False): 
            #EXCEL
            # Define variable to load the dataframe
            excel_dataframe = openpyxl.load_workbook("CDA_certifies-4-participantWorkshops.xlsx")

            # Define variable to read sheet
            dataframe = excel_dataframe.active

            # Cargar el archivo Excel
            wb = load_workbook('CDA_certifies-4-participantWorkshops.xlsx', read_only=True)

            # Obtener una lista de los nombres de las hojas
            sheet_names = wb.sheetnames

            # Iterar sobre cada hoja
            for sheet_name in sheet_names:
                # Obtener la hoja de trabajo por su nombre
                sheet = wb[sheet_name]
                
                # Iterar sobre las filas y columnas de la hoja
                for row in sheet.iter_rows(values_only=True):
                    _row = [row,]
                    for value in row:
                        _row.append(value)
                    
                    if(sheet_name=="Sheet1"):
                        
                        if(row[0]=="General Chair"):
                            Cargo1=_row[1]
                            Nombre1=_row[2]
                            Instituto1=_row[3]
                            Fecha_inicio=str(_row[5])+" "+str(_row[6])+"th"
                        elif(row[0]=="Jefa del Departamento de Investigación" or row[0]=="Director"):
                            Cargo2=_row[1]
                            Nombre2=_row[2]
                            Instituto2=_row[3]
                            Fecha_finalizacion=str(_row[5])+" "+str(_row[6])+"nd"
                            Year=_row[7]
                        elif (row[0]=="ID"):
                            print("Certifie")
                        else:
                            # INSTALAR LIVE SERVER, EJECUTAR SERVIDOR LOCAL Y EDITAR DIRECCION DE LAS IMAGENES LOCALES (TAMBIEN EN CSS)
                            
                            if _row[10]=='Si':
                                Participation=_row[5]
                                Title=_row[6]
                                Hours="of "+ str(_row[7])+" hours,"
                                image_url = server+"certifies.jpg"
                                width="250"
                                Hours="of "+ str(_row[7])+" hours,"
                                if (Participation=="Instructor"):
                                    Participacion="</br>imparted the course entitled"
                                elif (Participation=="Lecturer"):
                                    Participacion="</br>imparted an outstanding conference as Invited Speaker"
                                elif (Participation=="Invited Speaker"):
                                    Participacion="</br>imparted an outstanding conference as Invited Speaker"
                                elif (Participation=="Keynote Speaker"):  
                                    Participacion="</br>imparted an outstanding conference as Keynote Speaker"
                                elif (Participation=="Honorific advisor"):  
                                    Participacion="</br>imparted the conference entitled"
                                elif(Participation=="Workshop"):
                                    Participacion="</br>successfully completed the Workshop entitled"
                                elif(Participation=="Author"):
                                    Participacion="</br>performed the presentation of his Research Work identified with "
                                elif(Participation=="Staff"):
                                    Participacion="</br>participated as Staff"
                                elif(Participation=="Round Table"):
                                    Participacion="</br>participated as an outstanding member of the Round Table entitled "                  
                                elif(Participation=="Participant"):
                                    image_url = server+"certifiesthe.jpg"
                                    width="650"
                                Grado=_row[3]
                                NameAuthor=_row[2]
                                
                                # Crear el contenido HTML con la variable certifie-3
                                html_p1=f"""
                                <!DOCTYPE html>
                                <html lang="es">
                                <head>
                                <meta charset="UTF-8">
                                <meta http-equiv="X-UA-Compatible" content="IE=edge">
                                <meta name="viewport" content="width=device-width, initial-scale=1.0">  
                                <title>Certificado</title>
                                </head>
                                <body>
                            
                                    <div class="ex1">
                                        <p style="text-align: left; ">&nbsp;</p><p style="text-align: left;">&nbsp;</p>
                                        <p><img class="mi-img0" src="{server}upiita-logo.png" alt="UPIITA"  /><br/></p> 
                                        <p style="text-align: center;"><img src="{server}ICASST.png" alt="ICASST" width="130" height="130" />
                                        <img style="float: right;margin-right: 100px;" src="{server}ipn.jpg" alt="IPN" width="90" height="140" /></p>
                                        <p style="text-align: left;">&nbsp;</p>
                                        <div class="mi-texto0">The Scientific Committee of the </br>International Conference on AeroSpace Science and Technology </br></div>
                                        <p style="text-align: center; margin-top: 10mm;"><img src="{image_url}" alt="C" width={width} height="60" /></p> 
                                """
                                html_p2=""
                                if (Participation=="Lecturer" or Participation=="Invited Speaker" or Participation=="Keynote Speaker" or Participation=="Honorific advisor" 
                                    or Participation=="Author" or Participation=="Staff" or Participation=="Round Table" ):
                                    html_p2 = f"""
                                            <div class="mi-texto3"><br></br></div>
                                            <div class="mi-texto4">{NameAuthor}</br></div>
                                            <div class="mi-texto5">{Participacion}</br>{Title} at the</br></div>
                                            """
                                elif (Participation=="Instructor" or Participation=="Workshop") :
                                    html_p2 = f"""
                                    <div class="mi-texto3"></br>that </br></br></div>
                                    <div class="mi-texto4">{NameAuthor}</br></div>
                                    <div class="mi-texto5">{Participacion}</br>{Title} with a duration {Hours} at the</br></div>
                                    """
                                elif (Participation=="Participant"):
                                    html_p2 = f"""
                                    <div class="mi-texto3"></br>that </br></br></div>
                                    <div class="mi-texto4">{NameAuthor}</br></div>
                                    <div class="mi-texto5">{Participacion}</br>{Title} at the</br></div>
                                    """
                                
                                html_p3 = f"""
                                        <div class="mi-texto6"></br>International Conference on</br>AeroSpace Science and Technology (ICASST) {Year}</br></div>
                                        <div class="mi-texto7"></br>which took place from {Fecha_inicio} to</br>{Fecha_finalizacion} {Year}, at the facilities of UPIITA, IPN.</div>
                                        <br/>
                                        <br/>
                                """
                                if (Participation=="Participant" or Participation=="Workshop") :
                                    html_p4 = f""" 
                                            <table style="width: 100%; border-collapse: collapse; margin-left: 10px;" border="0">
                                            <tbody>
                                            <tr>
                                                <td style="width: 100%;"></td><td style="width: 100%;"><br /><br /></td>
                                            </tr>
                                            <tr>
                                            <td style="text-align: center;">
                                                <div class="mi-texto8" style="text-align: center; margin-left: 0px;">&nbsp;{Nombre1}<br />{Cargo1}<br/>{Instituto1}</div>
                                            </td>
                                            </tr>
                                            </tbody>
                                            </table>
                                        </div>
                                    </body>
                                    </html>
                                    """
                                else:
                                    html_p4 = f"""
                                        <table style="width: 100%; border-collapse: collapse; margin-left: 0px;" border="0">
                                        <tbody>
                                        <tr>
                                        <td style="width: 50%;"></td><td style="width: 50%;"><br /><br /></td>
                                        <td style="width: 50%;"></td><td style="width: 50%;"><br /><br /></td>
                                        </tr>
                                        <tr>
                                        <td style="width: 30%; text-align: center;">
                                            <div class="mi-texto8" style="text-align: center; margin-left: 0px;">&nbsp;{Nombre1}<br />{Cargo1}<br/>{Instituto1}</div>
                                        </td>
                                        <td style="width: 80%; text-align: center;">
                                            <div class="mi-texto8" style="text-align: center; margin-left: 0px;">&nbsp;{Nombre2}<br />{Cargo2}<br/>{Instituto2}</div>
                                        </td>
                                        </tr>
                                        
                                        </tbody>
                                        </table>
                                    </div>
                                </body>
                                </html>
                                """
                                    
                                html_content=html_p1+html_p2+html_p3+html_p4
                                # Guardar el HTML en un archivo
                                with open("certifie.html", "w", encoding="utf-8") as file:
                                    file.write(html_content)
                                if(NameAuthor!= None):
                                    pdf(NameAuthor)
                                    print("PDF "+NameAuthor+" ready") 
                                 # Archivos a manejar
            archivo_a_abrir = "C:\\Users\\tona2\\OneDrive\\Desktop\\Certifies_2024\\"+nameExcel
            # Abrir archivo cuando estamos haciendo pruebas
            print("Abriendo archivo "+nameExcel)
            open_excel_file(archivo_a_abrir)
        else:
            print("Excel "+nameExcel+" abierto, debes cerrarlo para poder continuar")

    else:
        print(f"El servidor respondió con un código de estado: {response.status_code}")
except requests.exceptions.RequestException as e:
    print("El servidor no está disponible.\nError:", e) 