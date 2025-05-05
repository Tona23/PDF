import jinja2
import pdfkit
import openpyxl
from openpyxl import load_workbook
import os
import requests
import win32com.client
import os

nameExcel="CDA_certifies-4-participantWorkshops.xlsx"

def open_excel_file(filepath):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    wb = excel.Workbooks.Open(os.path.abspath(filepath))
    return wb, excel

def obtener_archivos_excel_abiertos():
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        nombres = [wb.Name for wb in excel.Workbooks]
        return nombres
    except Exception as e:
        print("Excel "+nameExcel+" cerrado, el programa puede ejecutarse correctamente")
        return []

# Ejecutar
archivos_abiertos = obtener_archivos_excel_abiertos()

abierto=False
for nombre in archivos_abiertos:
    if nombre==nameExcel:
        abierto=True

# Limpiar consola
os.system('cls' if os.name == 'nt' else 'clear')

# Definir la URL de la imagen (puedes modificarla dinámicamente)
html_template = 'credencial.html'
server="http://127.0.0.1:5500/"
def pdf(nombre):

            options = {
                'page-width': '90.1mm',   # Ancho de página (carta en pulgadas)
                'page-height': '60.5mm',
            'margin-top': '0mm',   # Ajusta los márgenes superiores según sea necesario
                'margin-right': '0mm',  # Ajusta los márgenes derechos según sea necesario
                'margin-bottom': '0mm',  # Ajusta los márgenes inferiores según sea necesario
                'margin-left': '0mm',   'print-media-type': None ,
                }
            output_text =jinja2.Environment(loader=jinja2.FileSystemLoader(searchpath="./")).get_template(html_template).render("")

            config = pdfkit.configuration(wkhtmltopdf='wkhtmltopdf/bin/wkhtmltopdf.exe')
            output_pdf = 'Credencial '+nombre+'.pdf'
            pdfkit.from_string(output_text, output_pdf, configuration=config, css='credencial.css',options=options)

try:
    response = requests.get(server, timeout=5)  # timeout evita que se quede esperando demasiado
    if response.status_code == 200:

        if(abierto==False):
            #EXCEL
            # Define variable to load the dataframe
            excel_dataframe = openpyxl.load_workbook("CDA_certifies-4-participantWorkshops.xlsx")

            # Define variable to read sheet
            dataframe = excel_dataframe.active

            # Cargar el archivo Excel
            wb = load_workbook('CDA_certifies-4-participantWorkshops.xlsx', read_only=True)

            # Obtener una lista de los nombres de las hojas
            sheet_names = wb.sheetnames

            # Iterar sobre cada hoja
            for sheet_name in sheet_names:
                # Obtener la hoja de trabajo por su nombre
                sheet = wb[sheet_name]
                
                # Iterar sobre las filas y columnas de la hoja
                for row in sheet.iter_rows(values_only=True):
                    _row = [row,]
                    for value in row:
                        _row.append(value)
                    
                    if(sheet_name=="Sheet1"):
                        
                        if(row[0]=="General Chair"):
                        
                            MesInicio=_row[5][:3]
                            Fecha_inicio=_row[6]
                        elif(row[0]=="Encargado"):
                        
                            MesFinal=_row[5][:3]
                            Fecha_finalizacion=_row[6]
                            Year=_row[7]
                        elif (row[0]=="ID"):
                            print("Credencial")
                        else:
                            if(_row[13]=="Si"):

                                Instituto=_row[8]
                                Grado=_row[3]
                                Grado=Grado.replace("None", "")
                                Participation=_row[5]
                            
                                NameAuthor=_row[2]
                                if Participation=="Keynote Speaker":
                                    urlBarra= server+'speaker.png'
                                elif Participation=="Round Table":
                                    urlBarra= server+'round.png'
                                elif Participation=="Instructor":
                                    urlBarra= server+'instructor.png'
                                elif Participation=="Staff":
                                    urlBarra= server+'staff.png'
                                elif Participation=="Author":
                                    urlBarra= server+'author.png'
                                elif Participation=="Invited Speaker":
                                    urlBarra= server+'invited.jpeg'
                                elif Participation=="Participant":
                                    urlBarra= server+'participant.jpeg'
                                elif Participation=="Panelist":
                                    urlBarra= server+'panelist.png'
                                
                                html_content=f"""
                                <!DOCTYPE html>
                                <html lang="es">
                                <head>
                                    <meta charset="UTF-8">
                                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                </head>
                                <body>
                                    <div style="background-color: black;">
                                        <div class="contenedor-imagen">
                                                <img class="mi-img0" src="{server}upiita-logo.png" alt="UPIITA"  />
                                                <img class="mi-img" src="{server}International.jpeg" alt="International" /> 
                                                <img class="mi-img2" src="{server}2024.png" alt="2024" />
                                                <img class="mi-img1" src="{server}ICASST.png" alt="ICASST" />
                                                <img class="imagen-oscurecida" src="{server}fondoC.png" alt="Imagen oscurecida">
                                <img class="imagen" src="{urlBarra}" alt="Imagen oscurecida">
                                        </div>
                                    <div class="container">
                                    <p><br><br>{Grado}{NameAuthor}<p class="container0">{Instituto}</p>
                                    </div>
                                    <div class="contenedor-imagen">
                                        <img class="imagen1" src="{server}icasstweb.png" alt="Imagen oscurecida"/>
                                        <img class="imagen-oscurecida2" src="{server}fondoC.png" alt="Imagen oscurecida">
                                    </div>
                                    </div>
                                </div>
                                </body>
                                </html>
                                """;
                                
                                # Guardar el HTML en un archivo                                

                                with open("credencial.html", "w", encoding="utf-8") as file:
                                    file.write(html_content)
                                if(NameAuthor!= None):
                                    pdf(NameAuthor)
                                    print("PDF "+NameAuthor+" ready") 

            # Archivos a manejar
            archivo_a_abrir = "C:\\Users\\tona2\\OneDrive\\Desktop\\Certifies_2024\\"+nameExcel
            # Abrir archivo cuando estamos haciendo pruebas
            print("Abriendo archivo "+nameExcel)
            open_excel_file(archivo_a_abrir)
        else:
            print("Excel "+nameExcel+" abierto, debes cerrarlo para poder continuar")        
    else:
        print(f"El servidor respondió con un código de estado: {response.status_code}")
except requests.exceptions.RequestException as e:
    print("El servidor no está disponible.\nError:", e)
